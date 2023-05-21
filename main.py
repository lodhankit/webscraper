from bs4 import BeautifulSoup
import requests
import re
import nltk
import os
import pandas as pd
import numpy as np
import csv
from openpyxl import Workbook,load_workbook


def count_syllable(word):
    if(word[-2:] == "es" or word[-2:] == "ed"):
        word = word.replace(word[-2:],"")
    count = 0
    vowels = {'a','e','i','o','u'}
    for char in word:
        if(char in vowels):
            count += 1
    return count

website_df = pd.read_csv('input.csv')
web_list = list(website_df.iloc[:114,1].values)
web_id = list(website_df.iloc[:114,0].values)
# collecting all stopwords of all file into a set
wb = Workbook()
ws = wb.active
header = ['URL_ID', 'URL', 'POSITIVE SCORE', 'NEGATIVE SCORE','POLARITY SCORE','SUBJECTIVITY SCORE','AVG SENTENCE LENGTH','PERCENTAGE OF COMPLEX WORDS','FOG INDEX','AVG NUMBER OF WORDS PER SENTENCE','COMPLEX WORD COUNT','WORD COUNT','SYLLABLE PER WORD','PERSONAL PRONOUNS','AVG WORD LENGTH']
ws.append(header)
wb.save('Output Data Structure.xlsx')


stop_words = set()
stop_words_files = os.listdir('./StopWords')
for loc in stop_words_files:
    location = './StopWords/'+loc
    with open(location,'r') as file:
        stopword = file.read()
        token_stopword = re.sub(r"\n"," ",stopword)
        token_stopword = nltk.word_tokenize(stopword)
        new_words = [word.lower() for word in token_stopword if word.isalpha()]
        for w in new_words:
            stop_words.add(w)

# collecting all negetive words            
neg_words = set()
with open('./MasterDictionary/negative-words.txt','r') as file:
    negword = file.read()
    token_negword = re.sub(r"\n"," ",negword)
    token_negword = nltk.word_tokenize(negword)
    new_words = [word.lower() for word in token_negword if word.isalpha()]
    for w in new_words:
        neg_words.add(w)

# collecting all positive words            
pos_words = set()
with open('./MasterDictionary/positive-words.txt','r') as file:
    posword = file.read()
    token_posword = re.sub(r"\n"," ",posword)
    token_posword = nltk.word_tokenize(posword)
    new_words = [word.lower() for word in token_posword if word.isalpha()]
    for w in new_words:
        pos_words.add(w)
# extract urls and apply processing
i = 0
for loc in web_list:
    html_text = requests.get(loc).text
    content = BeautifulSoup(html_text,'lxml')
    try:
        heading = content.find('h1',class_='entry-title').text
        body = content.find('div',class_='td-post-content tagdiv-type').text
        body = re.sub(r"[\n\t]*","",body)
    
    except:
        try:
            heading = content.find('h1',class_='tdb-title-text').text
            body = content.find('div',class_='td_block_wrap tdb_single_content tdi_130 td-pb-border-top td_block_template_1 td-post-content tagdiv-type').text
            body = re.sub(r"[\n\t]*","",body)
    
        except:
            heading = None
            body = None
    
    txt_location = './extracted_files/'+str(int(web_id[i]))+'.txt'
    with open(txt_location,'w',encoding="utf-8") as file:
        file.write(f'Heading: {heading}\n')
        file.write(f'body: {body}')
    if(heading == None or body == None):
        data = [web_id[i], loc, 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN', 'NaN','NaN']
        ws.append(data)
        wb.save('Output Data Structure.xlsx')
        print(f'{web_id[i]} is saved.....')
        i+=1
        continue
    string = heading + " " + body
    tokens_with_US = nltk.word_tokenize(string)
    tokens = [word for word in tokens_with_US if word.isalpha()] # tokens that contains only words
    tokens_without_US = [word for word in tokens if word != "US"]
    tokens_lower = [word.lower() for word in tokens_without_US]
    tokens_filter = [] # append the word into this list that is not stop word
    for word in tokens_lower:
        if(word not in stop_words):
            tokens_filter.append(word)
    dic = {} # dictionary for positive and negetive counting
    dic['neg'] = 0
    dic['pos'] = 0
    for word in tokens_filter:
        if(word in neg_words):
            dic['neg'] += 1
        elif(word in pos_words):
            dic['pos'] += 1
        else:
            continue
    
    syllable_per_word = {}
    number_of_complex_words = 0
    for word in tokens:
        syllable_number = count_syllable(word.lower())
        syllable_per_word[word.lower()] = syllable_number
        if(syllable_number>2):
            number_of_complex_words += 1
    
    personal_words = {"i","we","my","ours","us"}
    personal_pronouns = 0
    for word in tokens_lower:
        if(word in personal_words):
            personal_pronouns += 1
    
    average_word_length = 0
    for word in tokens:
        average_word_length += len(word)
    average_word_length = average_word_length/len(tokens)
    
    
    polarity_score = (dic['pos']-dic['neg'])/((dic['pos']+dic['neg'])+0.000001)
    subjectivity_score = (dic['pos']+dic['neg'])/(len(tokens_filter)+0.000001)
    
    sentences = string.split('.')
    average_sentence_length = (len(tokens))/(len(sentences))
    percentage_of_complex_words = (number_of_complex_words)/(len(tokens))
    fog_index = 0.4*(average_sentence_length/percentage_of_complex_words)
    
    syllable_per_word = str(syllable_per_word)
    #wb = Workbook()
    #ws = wb.active
    data = [web_id[i],loc,dic['pos'], dic['neg'], polarity_score, subjectivity_score, average_sentence_length, percentage_of_complex_words, fog_index, average_sentence_length, number_of_complex_words, len(tokens), syllable_per_word, personal_pronouns, average_word_length]
    ws.append(data)
    wb.save('Output Data Structure.xlsx')
    print(f'{web_id[i]} is saved.....')
    i += 1